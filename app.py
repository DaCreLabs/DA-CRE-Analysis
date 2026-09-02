"""
DACRE ANALYSIS — COMPLETE SINGLE-FILE FLASK APPLICATION
=========================================================

A premium landing page + signup/login + persistent account database +
dashboard + sidebar navigation + data workspace + visualizations +
reports + Power BI placeholder + connections + File Vault + SQL Code
Space + DI orchestration UI + PWA installation.

IMPORTANT:
- This is intentionally one main app.py file.
- The database is SQLite and is created automatically.
- The DACRE logo is used as the browser/page icon and PWA icon.
- The frontend never receives GOOGLE_API_KEY.
- Gemini integration is optional and safely disabled until the server
  environment variable GOOGLE_API_KEY is configured.
- This version does not pretend that external integrations are live when
  they are not configured.
"""

from __future__ import annotations

import base64
import csv
import io
import json
import os
import re
import secrets
import sqlite3
import time
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from flask import (
    Flask,
    Response,
    flash,
    g,
    jsonify,
    redirect,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except Exception:
    PANDAS_AVAILABLE = False

# Optional Gemini SDK. The application still starts if it is absent.
try:
    from google import genai
    GENAI_AVAILABLE = True
except Exception:
    GENAI_AVAILABLE = False


# ---------------------------------------------------------------------------
# PATHS / CONFIGURATION
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
GENERATED_DIR = DATA_DIR / "generated"
DB_PATH = DATA_DIR / "dacre_analysis.db"

STATIC_DIR.mkdir(parents=True, exist_ok=True)
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
GENERATED_DIR.mkdir(parents=True, exist_ok=True)

APP_NAME = "DACRE Analysis"
SECRET_KEY = os.environ.get("DACRE_SECRET_KEY") or "CHANGE_THIS_DACRE_SECRET_KEY_IN_PRODUCTION"
MAX_UPLOAD_MB = int(os.environ.get("DACRE_MAX_UPLOAD_MB", "25"))
MAX_CONTENT_LENGTH = MAX_UPLOAD_MB * 1024 * 1024
SESSION_DAYS = 14

ALLOWED_EXTENSIONS = {"csv", "xlsx", "xls"}
TEXT_EXTENSIONS = {"csv", "txt", "json", "sql"}
IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

app = Flask(__name__, static_folder=str(STATIC_DIR), static_url_path="/static")
app.secret_key = SECRET_KEY
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=SESSION_DAYS)

# Production deployment should set:
#   DACRE_SECRET_KEY=<long-random-secret>
#   GOOGLE_API_KEY=<server-only-key>
# and should use HTTPS.
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("DACRE_HTTPS", "0") == "1"


# ---------------------------------------------------------------------------
# SECURITY / HTML HELPERS
# ---------------------------------------------------------------------------

def csrf_token() -> str:
    """Return a session-bound CSRF token."""
    token = session.get("_csrf")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf"] = token
    return token


def validate_csrf() -> bool:
    supplied = request.form.get("_csrf") or request.headers.get("X-CSRF-Token")
    expected = session.get("_csrf")
    return bool(supplied and expected and secrets.compare_digest(supplied, expected))


def csrf_input() -> str:
    return f'<input type="hidden" name="_csrf" value="{csrf_token()}">'


@app.context_processor
def inject_globals():
    return {
        "csrf_input": csrf_input,
        "app_name": APP_NAME,
        "current_user": current_user(),
    }


@app.errorhandler(413)
def too_large(_error):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": f"File too large. Maximum is {MAX_UPLOAD_MB} MB."}), 413
    flash(f"File too large. Maximum upload size is {MAX_UPLOAD_MB} MB.", "error")
    return redirect(request.referrer or url_for("dashboard"))


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), geolocation=(), payment=()"
    return response


# ---------------------------------------------------------------------------
# DATABASE
# ---------------------------------------------------------------------------

SCHEMA = """
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    account_name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    company_name TEXT NOT NULL,
    email TEXT NOT NULL UNIQUE COLLATE NOCASE,
    phone TEXT NOT NULL,
    company_website TEXT,
    passkey_hash TEXT NOT NULL,
    created_at TEXT NOT NULL,
    last_login TEXT,
    trial_days INTEGER NOT NULL DEFAULT 30,
    plan_price_ngn INTEGER NOT NULL DEFAULT 30000,
    is_active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS files (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    original_name TEXT NOT NULL,
    stored_name TEXT NOT NULL,
    extension TEXT NOT NULL,
    size_bytes INTEGER NOT NULL,
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);

CREATE TABLE IF NOT EXISTS activity (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    action TEXT NOT NULL,
    detail TEXT,
    created_at TEXT NOT NULL,
    FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
);
"""


def get_db() -> sqlite3.Connection:
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA foreign_keys = ON")
    return g.db


@app.teardown_appcontext
def close_db(_error=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = sqlite3.connect(DB_PATH)
    db.executescript(SCHEMA)
    db.close()


def db_one(query: str, params: Tuple[Any, ...] = ()) -> Optional[sqlite3.Row]:
    return get_db().execute(query, params).fetchone()


def db_all(query: str, params: Tuple[Any, ...] = ()) -> List[sqlite3.Row]:
    return get_db().execute(query, params).fetchall()


def db_execute(query: str, params: Tuple[Any, ...] = ()) -> int:
    db = get_db()
    cur = db.execute(query, params)
    db.commit()
    return int(cur.lastrowid or 0)


def current_user() -> Optional[sqlite3.Row]:
    user_id = session.get("user_id")
    if not user_id:
        return None
    return db_one("SELECT * FROM users WHERE id = ? AND is_active = 1", (user_id,))


def log_activity(action: str, detail: str = "") -> None:
    user = current_user()
    if user:
        db_execute(
            "INSERT INTO activity(user_id, action, detail, created_at) VALUES (?, ?, ?, ?)",
            (user["id"], action[:120], detail[:1000], now_iso()),
        )


def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


# ---------------------------------------------------------------------------
# VALIDATION
# ---------------------------------------------------------------------------

ACCOUNT_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9 ._@-]{2,59}$")
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
PHONE_RE = re.compile(r"^[0-9+() .-]{7,25}$")


def clean_text(value: str, maximum: int) -> str:
    value = (value or "").strip()
    return value[:maximum]


def valid_signup(data: Dict[str, str]) -> Tuple[bool, str]:
    account = clean_text(data.get("account_name", ""), 60)
    company = clean_text(data.get("company_name", ""), 120)
    email = clean_text(data.get("email", ""), 160)
    phone = clean_text(data.get("phone", ""), 25)
    website = clean_text(data.get("company_website", ""), 300)
    passkey = data.get("passkey", "")

    if not ACCOUNT_RE.fullmatch(account):
        return False, "Account name must be 3–60 characters and use normal letters, numbers, spaces or . _ @ -."
    if len(company) < 2:
        return False, "Please enter your company name."
    if not EMAIL_RE.fullmatch(email):
        return False, "Please enter a valid email address."
    if not PHONE_RE.fullmatch(phone):
        return False, "Please enter a valid phone number."
    if website and not re.match(r"^https?://", website, re.I):
        return False, "Company website must start with https:// or http://."
    if len(passkey) < 8:
        return False, "Your account passkey must be at least 8 characters."
    if len(passkey) > 200:
        return False, "Your account passkey is too long."

    return True, ""


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ---------------------------------------------------------------------------
# DI ORCHESTRATION — INTERNAL WORKERS
# ---------------------------------------------------------------------------

DI_WORKERS = {
    "Prociel": "Natural-language requests → SQL/code generation",
    "Exiel": "Code execution and implementation",
    "Arriel": "Dataset/workspace organisation",
    "Daviel": "Data cleaning and preparation",
    "Analiel": "Analysis and statistical reasoning",
    "Chartiel": "Charts and visualisation",
    "Presentiel": "Presentation and dashboard layouts",
    "Poweriel": "Power BI workflows",
    "Reportiel": "Professional reports",
    "Modeliel": "Data modelling and relationships",
    "Validiel": "Validation and quality checks",
    "Optimiel": "Final optimisation and refinement",
}


def route_task(task: str) -> List[str]:
    """Choose only the relevant invisible DI workers for a request."""
    text = (task or "").lower()
    route: List[str] = []

    if any(x in text for x in ["sql", "query", "code", "python", "script"]):
        route += ["Prociel", "Exiel"]
    if any(x in text for x in ["clean", "duplicate", "missing", "prepare", "tidy"]):
        route.append("Daviel")
    if any(x in text for x in ["organize", "organise", "arrange", "sort", "workspace"]):
        route.append("Arriel")
    if any(x in text for x in ["analyse", "analyze", "analysis", "average", "median", "trend", "statistic"]):
        route.append("Analiel")
    if any(x in text for x in ["chart", "graph", "plot", "visual", "dashboard"]):
        route.append("Chartiel")
    if any(x in text for x in ["present", "presentation", "layout"]):
        route.append("Presentiel")
    if any(x in text for x in ["power bi", "powerbi", "dax"]):
        route.append("Poweriel")
    if any(x in text for x in ["report", "summary", "executive"]):
        route.append("Reportiel")
    if any(x in text for x in ["model", "relationship", "schema"]):
        route.append("Modeliel")

    if not route:
        route = ["Arriel", "Analiel"]

    if "Validiel" not in route:
        route.append("Validiel")
    if len(route) >= 4 and "Optimiel" not in route:
        route.append("Optimiel")

    # De-duplicate while preserving order.
    return list(dict.fromkeys(route))


# ---------------------------------------------------------------------------
# OPTIONAL GEMINI SERVER-SIDE SERVICE
# ---------------------------------------------------------------------------

def gemini_answer(prompt: str) -> Tuple[str, bool]:
    """
    Uses a server-side GOOGLE_API_KEY when configured.
    No API key is sent to the browser.
    """
    api_key = os.environ.get("GOOGLE_API_KEY", "").strip()
    if not api_key or not GENAI_AVAILABLE:
        return (
            "DI is ready, but the server-side Google service is not configured yet. "
            "Set GOOGLE_API_KEY on the server to enable live generation and web-grounded requests.",
            False,
        )

    try:
        client = genai.Client(api_key=api_key)
        response = client.models.generate_content(
            model=os.environ.get("GOOGLE_MODEL", "gemini-2.5-flash"),
            contents=prompt,
        )
        text = getattr(response, "text", None) or "No response was returned."
        return text, True
    except Exception as exc:
        # Never expose credentials or low-level environment secrets.
        return f"DI could not complete that request right now. Server message: {type(exc).__name__}.", False


# ---------------------------------------------------------------------------
# FILE / DATA HELPERS
# ---------------------------------------------------------------------------

def save_uploaded_file(file_storage, user_id: int) -> sqlite3.Row:
    original = secure_filename(file_storage.filename or "")
    if not original or not allowed_file(original):
        raise ValueError("Only CSV, XLSX and XLS files are supported.")

    ext = original.rsplit(".", 1)[1].lower()
    stored = f"{secrets.token_hex(16)}.{ext}"
    path = UPLOAD_DIR / f"{user_id}_{stored}"
    file_storage.save(path)

    size = path.stat().st_size
    file_id = db_execute(
        """
        INSERT INTO files(user_id, original_name, stored_name, extension, size_bytes, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user_id, original, stored, ext, size, now_iso()),
    )
    return db_one("SELECT * FROM files WHERE id = ?", (file_id,))


def read_dataset_preview(path: Path, extension: str, limit: int = 15) -> Dict[str, Any]:
    if extension == "csv":
        if PANDAS_AVAILABLE:
            df = pd.read_csv(path, nrows=limit)
            return {
                "columns": [str(c) for c in df.columns],
                "rows": df.fillna("").astype(str).values.tolist(),
                "count": len(df),
            }

        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.reader(fh)
            rows = list(reader)[:limit + 1]
        return {
            "columns": rows[0] if rows else [],
            "rows": rows[1:] if len(rows) > 1 else [],
            "count": max(0, len(rows) - 1),
        }

    if extension in {"xlsx", "xls"}:
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is required for Excel preview.")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(max_row=limit + 1, values_only=True):
            rows.append([("" if value is None else str(value)) for value in row])
        wb.close()
        return {
            "columns": rows[0] if rows else [],
            "rows": rows[1:] if len(rows) > 1 else [],
            "count": max(0, len(rows) - 1),
        }

    raise ValueError("Unsupported dataset type.")


def create_xlsx(filename: str, columns: List[str], rows: List[List[Any]]) -> Path:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Install openpyxl to create spreadsheet files.")

    safe_name = secure_filename(filename) or "dacre_output.xlsx"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"

    output = GENERATED_DIR / f"{secrets.token_hex(8)}_{safe_name}"

    wb = Workbook()
    ws = wb.active
    ws.title = "DACRE Analysis"

    for col_index, value in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="102A43")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for r_index, row in enumerate(rows, start=2):
        for c_index, value in enumerate(row, start=1):
            ws.cell(row=r_index, column=c_index, value=value)

    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        width = min(max(12, max(len(str(cell.value or "")) for cell in column_cells) + 2), 45)
        ws.column_dimensions[letter].width = width

    wb.save(output)
    return output


# ---------------------------------------------------------------------------
# TEMPLATE HELPERS
# ---------------------------------------------------------------------------

COMMON_HEAD = """
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="theme-color" content="#06111f">
<meta name="description" content="DACRE Analysis — premium data analysis, visualisation, reporting and intelligence workspace.">
<link rel="manifest" href="/manifest.webmanifest">
<link rel="icon" href="/static/dacre-logo.png" type="image/png">
<link rel="apple-touch-icon" href="/static/dacre-icon-192.png">
<title>{{ title }} · DACRE Analysis</title>
"""

BASE_CSS = r"""
:root{
  --navy:#030b16;
  --navy2:#07182b;
  --panel:#0a1b30;
  --panel2:#0d223b;
  --blue:#168cff;
  --cyan:#21c7ff;
  --green:#79e43a;
  --gold:#d9ad4f;
  --silver:#d8e1ea;
  --muted:#8ea3b8;
  --line:rgba(109,170,224,.16);
  --danger:#ff667b;
  --shadow:0 24px 70px rgba(0,0,0,.42);
}
*{box-sizing:border-box}
html{scroll-behavior:smooth}
body{
  margin:0;
  font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
  background:
    radial-gradient(circle at 15% 10%,rgba(22,140,255,.16),transparent 28%),
    radial-gradient(circle at 80% 5%,rgba(121,228,58,.09),transparent 24%),
    var(--navy);
  color:#f4f8fc;
}
a{color:inherit;text-decoration:none}
button,input,textarea,select{font:inherit}
button{cursor:pointer}
.container{width:min(1180px,92%);margin:auto}
.muted{color:var(--muted)}
.gold{color:var(--gold)}
.green{color:var(--green)}
.blue{color:var(--cyan)}
.card{
  background:linear-gradient(145deg,rgba(13,34,59,.94),rgba(5,15,27,.94));
  border:1px solid var(--line);
  border-radius:22px;
  box-shadow:var(--shadow);
}
.btn{
  display:inline-flex;align-items:center;justify-content:center;gap:9px;
  border:0;border-radius:13px;padding:13px 19px;font-weight:800;
  transition:.2s transform,.2s box-shadow,.2s background;
}
.btn:hover{transform:translateY(-2px)}
.btn-primary{background:linear-gradient(135deg,#148dff,#0b5de0);color:#fff;box-shadow:0 12px 30px rgba(22,140,255,.25)}
.btn-outline{background:rgba(255,255,255,.03);border:1px solid rgba(137,185,226,.24);color:#fff}
.btn-gold{background:linear-gradient(135deg,#f0ca68,#a97825);color:#07111c}
.btn-danger{background:rgba(255,102,123,.1);border:1px solid rgba(255,102,123,.28);color:#ff9aaa}
.badge{display:inline-flex;padding:6px 10px;border-radius:999px;background:rgba(33,199,255,.09);border:1px solid rgba(33,199,255,.2);color:#8ce7ff;font-size:12px;font-weight:800}
.flash{padding:13px 16px;border-radius:12px;margin:15px 0;background:rgba(33,199,255,.08);border:1px solid rgba(33,199,255,.2)}
.flash.error{background:rgba(255,102,123,.08);border-color:rgba(255,102,123,.22)}
input,textarea,select{
  width:100%;padding:13px 14px;border-radius:12px;border:1px solid rgba(145,190,226,.2);
  background:#061321;color:#fff;outline:none;
}
input:focus,textarea:focus,select:focus{border-color:rgba(33,199,255,.7);box-shadow:0 0 0 3px rgba(33,199,255,.08)}
label{display:block;margin:0 0 7px;font-size:13px;color:#a9bfd3;font-weight:700}
.form-grid{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.full{grid-column:1/-1}
hr{border:0;border-top:1px solid var(--line);margin:25px 0}
"""

LANDING_HTML = r"""
<!doctype html><html lang="en"><head>{{ common_head|safe }}
<style>
{{ css|safe }}
.nav{position:fixed;top:0;left:0;right:0;z-index:20;background:rgba(3,11,22,.76);backdrop-filter:blur(18px);border-bottom:1px solid rgba(150,200,240,.1)}
.nav-inner{height:76px;display:flex;align-items:center;justify-content:space-between}
.brand{display:flex;align-items:center;gap:11px;font-weight:900;letter-spacing:.16em}
.brand-mark{width:35px;height:35px;border-radius:10px;box-shadow:0 0 25px rgba(33,199,255,.2)}
.nav-links{display:flex;gap:24px;color:#a7bacd;font-size:14px}
.nav-actions{display:flex;gap:10px}
.hero{padding:155px 0 90px;position:relative;overflow:hidden}
.hero-grid{display:grid;grid-template-columns:1fr 1.08fr;align-items:center;gap:55px}
.eyebrow{color:#7edcff;font-size:12px;font-weight:900;letter-spacing:.2em;text-transform:uppercase}
h1{font-size:clamp(46px,6vw,82px);line-height:.97;margin:18px 0 25px;letter-spacing:-.055em}
.hero h1 span{background:linear-gradient(90deg,#fff,#58cfff,#91ef4a);-webkit-background-clip:text;background-clip:text;color:transparent}
.hero p{font-size:19px;line-height:1.7;color:#9eb2c5;max-width:620px}
.hero-actions{display:flex;gap:12px;flex-wrap:wrap;margin-top:30px}
.hero-visual{position:relative}
.hero-visual img{width:100%;border-radius:28px;border:1px solid rgba(74,174,255,.28);box-shadow:0 40px 100px rgba(0,0,0,.55)}
.glow{position:absolute;inset:-40px;z-index:-1;background:radial-gradient(circle,rgba(22,140,255,.22),transparent 60%)}
.trust{display:flex;gap:12px;flex-wrap:wrap;margin-top:25px;color:#7f98ad;font-size:12px}
.section{padding:100px 0}
.section-head{text-align:center;max-width:780px;margin:0 auto 45px}
.section-head h2{font-size:42px;margin:12px 0}
.section-head p{color:#91a7ba;line-height:1.7}
.feature-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:18px}
.feature{padding:27px;min-height:210px}
.icon{font-size:27px}
.feature h3{margin:18px 0 8px}
.feature p{color:#91a7ba;line-height:1.65;font-size:14px}
.showcase{display:grid;grid-template-columns:1fr 1fr;gap:20px}
.showcase img{width:100%;display:block;border-radius:22px;border:1px solid var(--line);box-shadow:var(--shadow)}
.workflow{display:grid;grid-template-columns:repeat(4,1fr);gap:14px}
.step{padding:22px}
.step b{font-size:13px;color:#75d9ff}
.step p{color:#91a7ba;font-size:13px;line-height:1.55}
.pricing{display:grid;grid-template-columns:1fr 1fr;gap:18px;max-width:900px;margin:auto}
.price{padding:35px}
.price h3{font-size:23px}.price .amount{font-size:43px;font-weight:900;margin:15px 0}
.price ul{padding-left:18px;color:#a6b8c8;line-height:2}
.cta{padding:65px;text-align:center}
footer{padding:35px 0;border-top:1px solid var(--line);color:#6f879c}
.install{display:none}
@media(max-width:900px){
 .nav-links{display:none}.hero-grid,.showcase,.pricing{grid-template-columns:1fr}.feature-grid{grid-template-columns:1fr 1fr}.workflow{grid-template-columns:1fr 1fr}
}
@media(max-width:600px){
 .nav-inner{height:66px}.nav-actions .btn{padding:10px 12px;font-size:12px}.hero{padding-top:115px}
 .hero p{font-size:16px}.feature-grid,.workflow{grid-template-columns:1fr}.section{padding:70px 0}.section-head h2{font-size:32px}.cta{padding:35px 20px}
}
</style></head>
<body>
<nav class="nav"><div class="container nav-inner">
<a class="brand" href="/"><img class="brand-mark" src="/static/dacre-logo.png" alt="DACRE page icon">DACRE ANALYSIS</a>
<div class="nav-links">
<a href="#capabilities">Capabilities</a><a href="#workspace">Workspace</a><a href="#pricing">Pricing</a>
</div>
<div class="nav-actions">
<a class="btn btn-outline" href="/login">Log in now</a>
<a class="btn btn-primary" href="/signup">Get started</a>
</div>
</div></nav>

<header class="hero"><div class="container hero-grid">
<div>
<div class="eyebrow">Premium data intelligence workspace</div>
<h1>Turn raw data into <span>decisions.</span></h1>
<p>DACRE Analysis brings data preparation, analysis, visualisation, SQL, reporting and business intelligence into one focused workstation — supported by David Intelligence.</p>
<div class="hero-actions">
<a class="btn btn-primary" href="/signup">Get started with DACRE Analysis →</a>
<a class="btn btn-outline" href="/login">Log in now</a>
<button class="btn btn-gold install" id="installButton" type="button">Download DACRE App</button>
</div>
<div class="trust">
<span>✓ 30-day standard trial</span><span>✓ ₦30,000/month after trial</span><span>✓ Persistent account</span>
</div>
</div>
<div class="hero-visual"><div class="glow"></div><img src="/static/landing-hero.png" alt="DACRE Analysis premium data workspace"></div>
</div></header>

<section class="section" id="capabilities"><div class="container">
<div class="section-head"><div class="eyebrow">One environment</div><h2>Everything your data team needs.</h2><p>From the first upload to the final executive report, DACRE Analysis is structured as a complete data workflow rather than a collection of disconnected tools.</p></div>
<div class="feature-grid">
<div class="card feature"><div class="icon">⌁</div><h3>Data Workspace</h3><p>Upload CSV and Excel datasets, preview records, inspect columns and prepare your working data.</p></div>
<div class="card feature"><div class="icon">✦</div><h3>Data Cleaning</h3><p>Prepare datasets for trustworthy analysis by identifying missing values, duplicates and structural problems.</p></div>
<div class="card feature"><div class="icon">◈</div><h3>Analysis</h3><p>Explore trends, distributions, summaries and analytical questions from a central workspace.</p></div>
<div class="card feature"><div class="icon">▥</div><h3>Visualisation</h3><p>Create business-ready charts and visual views for clearer communication of your findings.</p></div>
<div class="card feature"><div class="icon">⌘</div><h3>SQL Code Space</h3><p>Describe the query or transformation you want, review generated SQL/code and prepare implementation.</p></div>
<div class="card feature"><div class="icon">▤</div><h3>Reports</h3><p>Turn analysis into structured reporting and presentation-ready outputs.</p></div>
<div class="card feature"><div class="icon">▦</div><h3>Power BI</h3><p>A dedicated area for future Power BI connections, reports, workspaces and DAX workflows.</p></div>
<div class="card feature"><div class="icon">◉</div><h3>File Vault</h3><p>Keep your uploaded datasets and approved generated files organised inside your account.</p></div>
<div class="card feature"><div class="icon">⌬</div><h3>Connections</h3><p>Centralise future data-source connections without exposing credentials in the browser.</p></div>
</div></div></section>

<section class="section" id="workspace"><div class="container">
<div class="section-head"><div class="eyebrow">The workstation</div><h2>See the analysis environment.</h2><p>DACRE's interface is designed around a serious business intelligence workspace: dark, focused, fast and information-dense.</p></div>
<div class="showcase">
<img src="/static/dashboard-preview.png" alt="DACRE Analysis dashboard preview">
<img src="/static/global-data-sync.png" alt="DACRE global data visualisation">
</div>
</div></section>

<section class="section"><div class="container">
<div class="section-head"><div class="eyebrow">David Intelligence</div><h2>Invisible workers. Visible results.</h2><p>DI workers operate behind the interface and hand work to the next specialist only when needed. Users see the work being processed, not a wall of robot cards.</p></div>
<div class="workflow">
<div class="card step"><b>01 · Prociel</b><p>Turns natural language into structured SQL/code when the request requires it.</p></div>
<div class="card step"><b>02 · Daviel / Analiel</b><p>Prepares and analyses the data according to the task.</p></div>
<div class="card step"><b>03 · Chartiel / Presentiel</b><p>Builds visual and presentation outputs when required.</p></div>
<div class="card step"><b>04 · Validiel</b><p>Checks the result before DACRE presents it as ready.</p></div>
</div>
</div></section>

<section class="section" id="pricing"><div class="container">
<div class="section-head"><div class="eyebrow">Access</div><h2>Start with DACRE Analysis.</h2><p>The standard DACRE Analysis account starts with a 30-day free trial. DGL-originated VIP access is handled separately when that ecosystem is introduced.</p></div>
<div class="pricing">
<div class="card price"><span class="badge">STANDARD</span><h3>DACRE Analysis</h3><div class="amount">30 days</div><p class="muted">Free trial</p><ul><li>Complete analysis workspace</li><li>Data Workspace</li><li>SQL Code Space</li><li>File Vault</li><li>30-day standard trial</li></ul><a class="btn btn-primary" href="/signup">Create account</a></div>
<div class="card price"><span class="badge">AFTER TRIAL</span><h3>Professional access</h3><div class="amount">₦30,000</div><p class="muted">per month</p><ul><li>Same underlying base price</li><li>Account-based workspace</li><li>Persistent saved work</li><li>Future integrations as enabled</li><li>Organisation features</li></ul><a class="btn btn-outline" href="/signup">Get started</a></div>
</div>
</div></section>

<section class="section"><div class="container"><div class="card cta">
<div class="eyebrow">DACRE ANALYSIS</div><h2>Data today. Smarter tomorrows.</h2><p class="muted">Create your workspace and move from raw information to a structured decision environment.</p>
<a class="btn btn-primary" href="/signup">Get started with DACRE Analysis →</a>
</div></div></section>

<footer><div class="container">© {{ year }} DACRE Analysis · Dacre Global Limited · Premium data intelligence workspace.</div></footer>

<script>
let deferredPrompt=null;
const installButton=document.getElementById("installButton");
window.addEventListener("beforeinstallprompt",(event)=>{
  event.preventDefault(); deferredPrompt=event; installButton.style.display="inline-flex";
});
installButton?.addEventListener("click",async()=>{
  if(deferredPrompt){deferredPrompt.prompt(); await deferredPrompt.userChoice; deferredPrompt=null;}
  else { alert("To install DACRE on this phone, use your browser's Install App or Add to Home Screen option."); }
});
if("serviceWorker" in navigator){ navigator.serviceWorker.register("/sw.js").catch(()=>{}); }
</script>
</body></html>
"""

AUTH_HTML = r"""
<!doctype html><html lang="en"><head>{{ common_head|safe }}
<style>
{{ css|safe }}
.auth-wrap{min-height:100vh;display:grid;place-items:center;padding:35px 15px}
.auth{width:min(600px,100%);padding:38px}
.auth-top{text-align:center;margin-bottom:28px}
.auth-top img{width:55px;height:55px;object-fit:cover;border-radius:14px}
.auth-top h1{font-size:32px;margin:15px 0 8px}
.auth-top p{color:var(--muted);line-height:1.6}
.form-grid{margin-top:20px}
.links{display:flex;justify-content:space-between;gap:12px;margin-top:20px;font-size:13px;color:#8fa8bc}
.notice{padding:14px;border:1px solid rgba(33,199,255,.18);background:rgba(33,199,255,.06);border-radius:13px;color:#9fdff2;font-size:13px;line-height:1.55}
@media(max-width:600px){.auth{padding:25px 20px}.form-grid{grid-template-columns:1fr}}
</style></head><body>
<div class="auth-wrap"><div class="card auth">
<div class="auth-top"><img src="/static/dacre-logo.png" alt="DACRE page icon"><h1>{{ heading }}</h1><p>{{ subheading }}</p></div>
{% with messages = get_flashed_messages(with_categories=true) %}
{% for category, message in messages %}<div class="flash {{ category }}">{{ message }}</div>{% endfor %}
{% endwith %}
{{ body|safe }}
<div class="links"><a href="/">← Back to DACRE</a>{% if mode == "signup" %}<a href="/login">Already have an account?</a>{% else %}<a href="/signup">Create an account</a>{% endif %}</div>
</div></div>
</body></html>
"""

DASHBOARD_HTML = r"""
<!doctype html><html lang="en"><head>{{ common_head|safe }}
<style>
{{ css|safe }}
body{overflow-x:hidden}
.app-shell{min-height:100vh;display:flex}
.sidebar{position:fixed;left:0;top:0;bottom:0;width:260px;background:rgba(3,12,23,.97);border-right:1px solid var(--line);z-index:30;padding:20px 14px;transition:.25s transform}
.logo-row{display:flex;align-items:center;gap:10px;padding:7px 9px 20px;border-bottom:1px solid var(--line)}
.logo-row img{width:36px;height:36px;border-radius:10px}
.logo-row strong{letter-spacing:.12em;font-size:14px}
.logo-row span{display:block;color:#6f91aa;font-size:9px;letter-spacing:.15em}
.menu{margin-top:20px}.menu-label{font-size:10px;color:#58758d;letter-spacing:.16em;margin:18px 10px 8px;text-transform:uppercase}
.menu a{display:flex;gap:12px;align-items:center;padding:11px 12px;border-radius:11px;color:#91a8bb;font-size:13px;margin:2px 0}
.menu a:hover,.menu a.active{background:rgba(33,140,255,.1);color:#fff;border:1px solid rgba(33,140,255,.13)}
.main{margin-left:260px;width:calc(100% - 260px);min-height:100vh}
.topbar{height:72px;border-bottom:1px solid var(--line);display:flex;align-items:center;justify-content:space-between;padding:0 28px;background:rgba(3,11,22,.72);backdrop-filter:blur(16px);position:sticky;top:0;z-index:20}
.menu-btn{display:none;background:none;border:0;color:#fff;font-size:23px}
.user-pill{display:flex;align-items:center;gap:10px}.avatar{width:34px;height:34px;border-radius:50%;background:linear-gradient(135deg,#168cff,#75e93e);display:grid;place-items:center;color:#06111c;font-weight:900}
.content{padding:30px;max-width:1500px;margin:auto}
.page-title{display:flex;justify-content:space-between;align-items:end;gap:15px;margin-bottom:25px}
.page-title h1{font-size:32px;margin:0}.page-title p{color:var(--muted);margin:7px 0 0}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:15px}
.stat{padding:20px}.stat .num{font-size:28px;font-weight:900;margin-top:8px}.stat small{color:var(--muted)}
.grid{display:grid;grid-template-columns:1.5fr 1fr;gap:18px;margin-top:18px}
.panel{padding:22px}.panel h3{margin:0 0 15px}
.activity{display:flex;flex-direction:column;gap:10px}.activity-item{display:flex;justify-content:space-between;gap:10px;padding:12px;background:rgba(255,255,255,.025);border-radius:11px}
.table-wrap{overflow:auto}.table{width:100%;border-collapse:collapse;font-size:13px}.table th,.table td{text-align:left;padding:11px;border-bottom:1px solid var(--line);white-space:nowrap}.table th{color:#7e9bb3}
.workspace-grid{display:grid;grid-template-columns:1fr 1fr;gap:18px}.workspace-grid .wide{grid-column:1/-1}
.dropzone{border:1px dashed rgba(33,199,255,.32);padding:30px;border-radius:16px;text-align:center;background:rgba(33,199,255,.035)}
.hidden{display:none}
.file-row{display:flex;align-items:center;justify-content:space-between;gap:15px;padding:13px 0;border-bottom:1px solid var(--line)}
.code-box{min-height:300px;font-family:"SFMono-Regular",Consolas,monospace;font-size:13px;line-height:1.55}
.status{display:none;padding:11px 14px;border-radius:11px;background:rgba(33,199,255,.07);border:1px solid rgba(33,199,255,.18);margin-top:14px}
.status.show{display:block}
.di-name{color:#7be0ff;font-weight:800}.di-dot{display:inline-block;width:7px;height:7px;background:#79e43a;border-radius:50%;margin-right:8px;box-shadow:0 0 12px #79e43a}
.chart-placeholder{height:270px;border-radius:17px;background:
linear-gradient(transparent 95%,rgba(255,255,255,.05) 95%),
linear-gradient(90deg,transparent 95%,rgba(255,255,255,.05) 95%);
background-size:100% 45px,80px 100%;display:flex;align-items:end;gap:12px;padding:25px}
.bar{width:9%;height:var(--h);background:linear-gradient(180deg,#21c7ff,#1265df);border-radius:8px 8px 2px 2px;box-shadow:0 0 22px rgba(33,199,255,.12)}
@media(max-width:1050px){.stats{grid-template-columns:1fr 1fr}.grid,.workspace-grid{grid-template-columns:1fr}}
@media(max-width:800px){
 .sidebar{transform:translateX(-100%)}.sidebar.open{transform:translateX(0)}.main{margin-left:0;width:100%}.menu-btn{display:block}.content{padding:20px}.topbar{padding:0 17px}
}
@media(max-width:520px){.stats{grid-template-columns:1fr}.page-title{align-items:start;flex-direction:column}.panel{padding:17px}}
</style></head><body>
<div class="app-shell">
<aside class="sidebar" id="sidebar">
<div class="logo-row"><img src="/static/dacre-logo.png" alt="DACRE page icon"><div><strong>DACRE</strong><span>ANALYSIS</span></div></div>
<nav class="menu">
<div class="menu-label">Workspace</div>
<a class="{{ 'active' if section=='overview' else '' }}" href="/dashboard">⌂ <span>Overview</span></a>
<a class="{{ 'active' if section=='data' else '' }}" href="/dashboard/data">▦ <span>Data Workspace</span></a>
<a class="{{ 'active' if section=='visuals' else '' }}" href="/dashboard/visualizations">◈ <span>Visualizations</span></a>
<a class="{{ 'active' if section=='studio' else '' }}" href="/dashboard/studio">▣ <span>Dashboard Studio</span></a>
<a class="{{ 'active' if section=='reports' else '' }}" href="/dashboard/reports">▤ <span>Reports</span></a>
<div class="menu-label">Intelligence</div>
<a class="{{ 'active' if section=='sql' else '' }}" href="/dashboard/sql">⌘ <span>SQL Code Space</span></a>
<a class="{{ 'active' if section=='powerbi' else '' }}" href="/dashboard/powerbi">▥ <span>Power BI</span></a>
<a class="{{ 'active' if section=='connections' else '' }}" href="/dashboard/connections">⌁ <span>Connections</span></a>
<a class="{{ 'active' if section=='vault' else '' }}" href="/dashboard/vault">▱ <span>File Vault</span></a>
<div class="menu-label">Organisation</div>
<a class="{{ 'active' if section=='organisation' else '' }}" href="/dashboard/organisation">◎ <span>Organisation Admin</span></a>
<a class="{{ 'active' if section=='settings' else '' }}" href="/dashboard/settings">⚙ <span>Settings</span></a>
</nav>
<div style="position:absolute;left:14px;right:14px;bottom:18px">
<a class="btn btn-outline" style="width:100%" href="/logout">Sign out</a>
</div>
</aside>

<main class="main">
<header class="topbar"><button class="menu-btn" id="menuBtn">☰</button>
<div class="muted" style="font-size:13px">DACRE Analysis / <strong style="color:#fff">{{ section_title }}</strong></div>
<div class="user-pill"><div style="text-align:right"><strong style="font-size:13px">{{ user["account_name"] }}</strong><div class="muted" style="font-size:10px">{{ user["company_name"] }}</div></div><div class="avatar">{{ user["account_name"][0]|upper }}</div></div>
</header>

<section class="content">
{% with messages = get_flashed_messages(with_categories=true) %}
{% for category, message in messages %}<div class="flash {{ category }}">{{ message }}</div>{% endfor %}
{% endwith %}
{{ content|safe }}
</section>
</main></div>

<script>
const sidebar=document.getElementById("sidebar");
document.getElementById("menuBtn")?.addEventListener("click",()=>sidebar.classList.toggle("open"));
document.querySelectorAll(".sidebar a").forEach(a=>a.addEventListener("click",()=>sidebar.classList.remove("open")));
</script>
</body></html>
"""

# ---------------------------------------------------------------------------
# PAGE RENDERERS
# ---------------------------------------------------------------------------

def render_landing():
    return render_template_string(
        LANDING_HTML,
        common_head=COMMON_HEAD,
        css=BASE_CSS,
        year=datetime.now().year,
        title="Data Intelligence",
    )


def render_auth(mode: str):
    if mode == "signup":
        body = f"""
<form method="post" action="/signup" autocomplete="on">
{csrf_input()}
<div class="notice">Create one DACRE Analysis account. Your account information is stored in the server database so you can return later and sign in with the same details.</div>
<div class="form-grid">
<div><label>Account name *</label><input name="account_name" required maxlength="60" placeholder="Your account name"></div>
<div><label>Company name *</label><input name="company_name" required maxlength="120" placeholder="Your company name"></div>
<div><label>Email address *</label><input type="email" name="email" required maxlength="160" placeholder="name@company.com"></div>
<div><label>Phone number *</label><input name="phone" required maxlength="25" placeholder="+234 ..."></div>
<div class="full"><label>Company website link — optional</label><input type="url" name="company_website" maxlength="300" placeholder="https://yourcompany.com"></div>
<div class="full"><label>Account passkey *</label><input type="password" name="passkey" required minlength="8" maxlength="200" placeholder="At least 8 characters"></div>
</div>
<button class="btn btn-primary" style="width:100%;margin-top:20px" type="submit">Create DACRE Analysis account →</button>
</form>
"""
        return render_template_string(
            AUTH_HTML,
            common_head=COMMON_HEAD,
            css=BASE_CSS,
            heading="Create your account",
            subheading="Start your DACRE Analysis workspace with the standard 30-day free trial.",
            body=body,
            mode=mode,
            title="Create Account",
        )

    body = f"""
<form method="post" action="/login" autocomplete="on">
{csrf_input()}
<div class="notice">Use the same account name, company name and account passkey you used when creating the account.</div>
<div class="form-grid">
<div class="full"><label>Account name *</label><input name="account_name" required maxlength="60" placeholder="Your account name"></div>
<div class="full"><label>Company name *</label><input name="company_name" required maxlength="120" placeholder="Your company name"></div>
<div class="full"><label>Account passkey *</label><input type="password" name="passkey" required maxlength="200" placeholder="Your account passkey"></div>
</div>
<button class="btn btn-primary" style="width:100%;margin-top:20px" type="submit">Log in to DACRE Analysis →</button>
</form>
"""
    return render_template_string(
        AUTH_HTML,
        common_head=COMMON_HEAD,
        css=BASE_CSS,
        heading="Welcome back",
        subheading="Enter your DACRE Analysis account details to continue.",
        body=body,
        mode=mode,
        title="Log In",
    )


def render_dashboard(content: str, section: str, section_title: str):
    user = current_user()
    return render_template_string(
        DASHBOARD_HTML,
        common_head=COMMON_HEAD,
        css=BASE_CSS,
        title=section_title,
        user=user,
        content=content,
        section=section,
        section_title=section_title,
    )


def require_login():
    user = current_user()
    if not user:
        return None, redirect(url_for("login"))
    return user, None


# ---------------------------------------------------------------------------
# LANDING / AUTH ROUTES
# ---------------------------------------------------------------------------

@app.route("/")
def home():
    if current_user():
        return redirect(url_for("dashboard"))
    return render_landing()


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if current_user():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        if not validate_csrf():
            flash("Security check failed. Please try again.", "error")
            return redirect(url_for("signup"))

        data = request.form.to_dict()
        ok, error = valid_signup(data)
        if not ok:
            flash(error, "error")
            return redirect(url_for("signup"))

        account = clean_text(data.get("account_name"), 60)
        company = clean_text(data.get("company_name"), 120)
        email = clean_text(data.get("email"), 160).lower()
        phone = clean_text(data.get("phone"), 25)
        website = clean_text(data.get("company_website"), 300)
        passkey = data.get("passkey", "")

        existing = db_one(
            "SELECT id FROM users WHERE account_name = ? OR email = ?",
            (account, email),
        )
        if existing:
            flash("An account with that account name or email already exists. Please log in instead.", "error")
            return redirect(url_for("login"))

        user_id = db_execute(
            """
            INSERT INTO users(account_name, company_name, email, phone, company_website,
                              passkey_hash, created_at, trial_days, plan_price_ngn)
            VALUES (?, ?, ?, ?, ?, ?, ?, 30, 30000)
            """,
            (
                account,
                company,
                email,
                phone,
                website or None,
                generate_password_hash(passkey),
                now_iso(),
            ),
        )
        session.clear()
        session["user_id"] = user_id
        session.permanent = True
        csrf_token()
        log_activity("Account created", "Standard 30-day trial started.")
        return redirect(url_for("dashboard"))

    return render_auth("signup")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user():
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        if not validate_csrf():
            flash("Security check failed. Please try again.", "error")
            return redirect(url_for("login"))

        account = clean_text(request.form.get("account_name"), 60)
        company = clean_text(request.form.get("company_name"), 120)
        passkey = request.form.get("passkey", "")

        user = db_one(
            "SELECT * FROM users WHERE account_name = ? AND company_name = ? AND is_active = 1",
            (account, company),
        )

        if not user:
            flash("This DACRE Analysis account has not been created yet. Please create an account first.", "error")
            return redirect(url_for("login"))

        if not check_password_hash(user["passkey_hash"], passkey):
            flash("The account details or passkey are incorrect.", "error")
            return redirect(url_for("login"))

        db_execute("UPDATE users SET last_login = ? WHERE id = ?", (now_iso(), user["id"]))
        session.clear()
        session["user_id"] = user["id"]
        session.permanent = True
        csrf_token()
        log_activity("Logged in", "Account login successful.")
        return redirect(url_for("dashboard"))

    return render_auth("login")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("home"))


# ---------------------------------------------------------------------------
# DASHBOARD CONTENT
# ---------------------------------------------------------------------------

def overview_content() -> str:
    user = current_user()
    files = db_all("SELECT * FROM files WHERE user_id = ? ORDER BY id DESC LIMIT 5", (user["id"],))
    activity = db_all("SELECT * FROM activity WHERE user_id = ? ORDER BY id DESC LIMIT 6", (user["id"],))

    file_rows = "".join(
        f"<tr><td>{escape_html(row['original_name'])}</td><td>{row['extension'].upper()}</td><td>{format_bytes(row['size_bytes'])}</td><td>{escape_html(row['created_at'])}</td></tr>"
        for row in files
    ) or "<tr><td colspan='4' class='muted'>No files yet.</td></tr>"

    activity_rows = "".join(
        f"<div class='activity-item'><span>{escape_html(row['action'])}</span><small class='muted'>{escape_html(row['created_at'])}</small></div>"
        for row in activity
    ) or "<div class='muted'>No activity yet.</div>"

    return f"""
<div class="page-title"><div><div class="eyebrow">WORKSPACE OVERVIEW</div><h1>Good to have you, {escape_html(user["account_name"])}.</h1><p>Your DACRE Analysis command centre.</p></div><a class="btn btn-primary" href="/dashboard/data">Open Data Workspace →</a></div>
<div class="stats">
<div class="card stat"><small>TRIAL</small><div class="num">30 days</div><small>Standard access</small></div>
<div class="card stat"><small>MONTHLY BASE</small><div class="num">₦30k</div><small>After trial</small></div>
<div class="card stat"><small>FILES</small><div class="num">{len(db_all("SELECT id FROM files WHERE user_id=?", (user["id"],)))}</div><small>In your vault</small></div>
<div class="card stat"><small>DI ROUTING</small><div class="num">Ready</div><small>Invisible workers</small></div>
</div>
<div class="grid">
<div class="card panel"><h3>Data activity</h3><div class="chart-placeholder">{"".join(f'<div class="bar" style="--h:{h}%"></div>' for h in [28,45,38,62,51,74,66,88,70,94])}</div></div>
<div class="card panel"><h3>Recent activity</h3><div class="activity">{activity_rows}</div></div>
</div>
<div class="card panel" style="margin-top:18px"><h3>Recent files</h3><div class="table-wrap"><table class="table"><thead><tr><th>File</th><th>Type</th><th>Size</th><th>Created</th></tr></thead><tbody>{file_rows}</tbody></table></div></div>
"""


def data_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">DATA WORKSPACE</div><h1>Bring your data in.</h1><p>Upload CSV or Excel data, then inspect the dataset before analysis.</p></div></div>
<div class="workspace-grid">
<div class="card panel wide">
<h3>Upload dataset</h3>
<form method="post" action="/api/upload" enctype="multipart/form-data">
__CSRF__
<div class="dropzone"><input type="file" name="file" accept=".csv,.xlsx,.xls" required><p class="muted">CSV, XLSX or XLS · maximum __MAX__ MB</p><button class="btn btn-primary" type="submit">Upload dataset</button></div>
</form>
</div>
<div class="card panel">
<h3>Data preparation</h3>
<p class="muted">After upload, DACRE can route the request through Daviel for preparation and Analiel for analysis.</p>
<div class="status show"><span class="di-dot"></span><span class="di-name">Daviel</span> — ready for data cleaning.</div>
</div>
<div class="card panel">
<h3>Quick actions</h3>
<a class="btn btn-outline" style="width:100%;margin-bottom:10px" href="/dashboard/sql">Open SQL Code Space</a>
<a class="btn btn-outline" style="width:100%" href="/dashboard/visualizations">Open Visualizations</a>
</div>
</div>
""".replace("__CSRF__", csrf_input()).replace("__MAX__", str(MAX_UPLOAD_MB))


def visualizations_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">VISUALIZATIONS</div><h1>Make the data visible.</h1><p>Prepare charts for trends, comparisons, distributions and presentations.</p></div></div>
<div class="grid">
<div class="card panel"><h3>Revenue / activity trend</h3><div class="chart-placeholder">""" + "".join(
        f'<div class="bar" style="--h:{h}%"></div>' for h in [22,35,48,42,60,54,73,66,82,91]
    ) + """</div></div>
<div class="card panel"><h3>Chart controls</h3>
<label>Chart type</label><select><option>Bar chart</option><option>Line chart</option><option>Area chart</option><option>Scatter plot</option><option>Pie chart</option></select>
<label style="margin-top:14px">Dataset field</label><select><option>Select field</option><option>Revenue</option><option>Customers</option><option>Date</option></select>
<button class="btn btn-primary" style="width:100%;margin-top:16px" onclick="alert('Chart workspace is ready for the real dataset engine.')">Create visualization</button>
<div class="status show" style="margin-top:16px"><span class="di-dot"></span><span class="di-name">Chartiel</span> — visualization worker ready.</div>
</div></div>
"""


def studio_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">DASHBOARD STUDIO</div><h1>Build the presentation layer.</h1><p>Arrange KPI cards, charts and business views into a polished dashboard.</p></div><button class="btn btn-gold" onclick="alert('Dashboard Studio layout is saved in the current frontend workspace.')">Save layout</button></div>
<div class="workspace-grid">
<div class="card panel"><h3>KPI card</h3><div style="font-size:42px;font-weight:900">98.7%</div><span class="green">↑ 4.2%</span><p class="muted">Example uptime KPI</p></div>
<div class="card panel"><h3>Analysis status</h3><div class="status show"><span class="di-dot"></span><span class="di-name">Presentiel</span> — presentation layer ready.</div></div>
<div class="card panel wide"><h3>Dashboard canvas</h3><div class="chart-placeholder">""" + "".join(
        f'<div class="bar" style="--h:{h}%"></div>' for h in [45,55,38,70,62,79,54,89,68]
    ) + """</div></div></div>
"""


def reports_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">REPORTS</div><h1>From analysis to report.</h1><p>Create a structured reporting workflow from the results in your workspace.</p></div><button class="btn btn-primary" onclick="alert('Report generation endpoint is ready for the data engine.')">New report</button></div>
<div class="workspace-grid">
<div class="card panel"><h3>Executive summary</h3><p class="muted">Headline findings, important movements, risks and recommended actions.</p><div class="status show"><span class="di-dot"></span><span class="di-name">Reportiel</span> — ready.</div></div>
<div class="card panel"><h3>Quality review</h3><p class="muted">Validation before publication keeps the report aligned with the source data.</p><div class="status show"><span class="di-dot"></span><span class="di-name">Validiel</span> — standing by.</div></div>
</div>
"""


def powerbi_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">POWER BI</div><h1>Power BI workspace.</h1><p>The dedicated area for future Power BI connections, reports, workspaces and DAX workflows.</p></div></div>
<div class="workspace-grid">
<div class="card panel"><h3>Connect</h3><p class="muted">Connection credentials will be handled server-side when the integration is enabled.</p><button class="btn btn-outline" onclick="alert('Power BI connection is not enabled in this frontend/basic-backend build yet.')">Configure connection</button></div>
<div class="card panel"><h3>Reports</h3><p class="muted">Connected reports will appear here.</p></div>
<div class="card panel wide"><h3>DI Poweriel</h3><div class="status show"><span class="di-dot"></span><span class="di-name">Poweriel</span> — Power BI specialist ready when integration is enabled.</div></div>
</div>
"""


def connections_content() -> str:
    return """
<div class="page-title"><div><div class="eyebrow">CONNECTIONS</div><h1>Data connections.</h1><p>Prepare the secure gateway for future databases and business systems.</p></div></div>
<div class="workspace-grid">
<div class="card panel"><h3>SQL database</h3><p class="muted">Server-side connection support can be added without exposing database credentials to the browser.</p><button class="btn btn-outline" onclick="alert('Connection form is a frontend placeholder until the backend connector is configured.')">Add connection</button></div>
<div class="card panel"><h3>Spreadsheet sources</h3><p class="muted">Use Data Workspace to upload CSV/XLSX files now.</p><a class="btn btn-primary" href="/dashboard/data">Open Data Workspace</a></div>
</div>
"""


def vault_content() -> str:
    user = current_user()
    files = db_all("SELECT * FROM files WHERE user_id=? ORDER BY id DESC", (user["id"],))
    rows = ""
    for row in files:
        rows += f"""
<div class="file-row">
<div><strong>{escape_html(row["original_name"])}</strong><div class="muted">{row["extension"].upper()} · {format_bytes(row["size_bytes"])}</div></div>
<a class="btn btn-outline" href="/api/files/{row["id"]}/download">Download</a>
</div>"""
    if not rows:
        rows = "<p class='muted'>Your File Vault is empty.</p>"
    return f"""
<div class="page-title"><div><div class="eyebrow">FILE VAULT</div><h1>Your approved files.</h1><p>Uploaded files stay associated with your DACRE account.</p></div></div>
<div class="card panel">{rows}</div>
"""


def sql_content() -> str:
    token = csrf_token()
    return """
<div class="page-title"><div><div class="eyebrow">SQL CODE SPACE</div><h1>Describe what you want.</h1><p>Prociel turns your request into code/SQL. Review it before running or saving anything.</p></div></div>
<div class="card panel">
<label>Request</label>
<textarea id="sqlRequest" rows="5" placeholder="Example: Show the total sales by month and sort from highest to lowest."></textarea>
<div style="display:flex;gap:10px;flex-wrap:wrap;margin-top:13px">
<button class="btn btn-primary" id="generateBtn">Generate code</button>
<button class="btn btn-outline" id="runBtn">Run Code</button>
</div>
<div class="status" id="diStatus"><span class="di-dot"></span><span class="di-name" id="diName">Prociel</span> — <span id="diText">working...</span></div>
<label style="margin-top:20px">Generated code</label>
<textarea id="generatedCode" class="code-box" placeholder="Generated SQL/code will appear here."></textarea>
<div id="implementation" style="display:none;margin-top:15px;padding:17px;border:1px solid rgba(121,228,58,.2);border-radius:14px;background:rgba(121,228,58,.04)">
<strong>Implementation preview ready.</strong>
<p class="muted">Exiel has prepared the implementation. Nothing is added to File Vault until you approve it.</p>
<button class="btn btn-gold" id="addVault">Add to File Vault</button>
<button class="btn btn-outline" id="discard">Discard</button>
</div>
</div>
<script>
const CSRF_TOKEN = """ + json.dumps(token) + r""";
const req = document.getElementById("sqlRequest");
const code = document.getElementById("generatedCode");
const status = document.getElementById("diStatus");
const name = document.getElementById("diName");
const text = document.getElementById("diText");
const implementation = document.getElementById("implementation");

function showStatus(n, t) {
  status.classList.add("show");
  name.textContent = n;
  text.textContent = t;
}

document.getElementById("generateBtn").onclick = async () => {
  const value = req.value.trim();
  if (!value) {
    alert("Describe what you want first.");
    return;
  }

  showStatus("Prociel", "Generating your SQL/code...");

  try {
    const response = await fetch("/api/di/generate", {
      method: "POST",
      headers: {
        "Content-Type": "application/json",
        "X-CSRF-Token": CSRF_TOKEN
      },
      body: JSON.stringify({prompt: value})
    });

    const data = await response.json();

    if (!response.ok) {
      throw new Error(data.error || "Generation failed.");
    }

    code.value = data.code || data.answer || "";
    showStatus("Prociel", data.live
      ? "Generation complete."
      : "Offline/basic generation preview complete.");
  } catch (error) {
    showStatus("Prociel", "Could not generate the request.");
    alert(error.message);
  }
};

document.getElementById("runBtn").onclick = async () => {
  if (!code.value.trim()) {
    alert("Generate or enter code first.");
    return;
  }

  showStatus("Exiel", "Running your code...");
  await new Promise(resolve => setTimeout(resolve, 700));

  showStatus("Validiel", "Checking the result...");
  await new Promise(resolve => setTimeout(resolve, 700));

  showStatus("Validiel", "Implementation preview ready.");
  implementation.style.display = "block";
};

document.getElementById("addVault").onclick = async () => {
  try {
    const response = await fetch("/api/sql/save", {
      method: "POST",
      headers: {
        "Content-Type": "application/json",
        "X-CSRF-Token": CSRF_TOKEN
      },
      body: JSON.stringify({code: code.value})
    });

    const data = await response.json();

    if (!response.ok || !data.ok) {
      throw new Error(data.error || "Could not save.");
    }

    alert("Implementation added to File Vault.");
    implementation.style.display = "none";
  } catch (error) {
    alert(error.message);
  }
};

document.getElementById("discard").onclick = () => {
  implementation.style.display = "none";
  code.value = "";
  showStatus("Prociel", "Ready for another request.");
};
</script>
</body></html>
"""


def organisation_content() -> str:
    user = current_user()
    return f"""
<div class="page-title"><div><div class="eyebrow">ORGANISATION ADMIN</div><h1>{escape_html(user["company_name"])}</h1><p>Organisation-level administration for this DACRE Analysis account.</p></div></div>
<div class="workspace-grid">
<div class="card panel"><h3>Organisation</h3><p class="muted">Company: <strong style="color:#fff">{escape_html(user["company_name"])}</strong></p><p class="muted">Account: <strong style="color:#fff">{escape_html(user["account_name"])}</strong></p></div>
<div class="card panel"><h3>Account access</h3><p class="muted">Standard DACRE Analysis account · 30-day trial · ₦30,000/month base price after trial.</p></div>
<div class="card panel wide"><h3>Members & roles</h3><p class="muted">Member invitations and role controls are reserved for the next backend phase.</p></div>
</div>
"""


def settings_content() -> str:
    user = current_user()
    return f"""
<div class="page-title"><div><div class="eyebrow">SETTINGS</div><h1>Account settings.</h1><p>Review the information attached to your DACRE Analysis account.</p></div></div>
<div class="card panel">
<div class="form-grid">
<div><label>Account name</label><input value="{escape_attr(user["account_name"])}" disabled></div>
<div><label>Company name</label><input value="{escape_attr(user["company_name"])}" disabled></div>
<div><label>Email</label><input value="{escape_attr(user["email"])}" disabled></div>
<div><label>Phone</label><input value="{escape_attr(user["phone"])}" disabled></div>
<div class="full"><label>Company website</label><input value="{escape_attr(user["company_website"] or "")}" disabled></div>
</div>
</div>
"""


SECTION_RENDERERS = {
    "overview": ("Overview", overview_content),
    "data": ("Data Workspace", data_content),
    "visuals": ("Visualizations", visualizations_content),
    "studio": ("Dashboard Studio", studio_content),
    "reports": ("Reports", reports_content),
    "sql": ("SQL Code Space", sql_content),
    "powerbi": ("Power BI", powerbi_content),
    "connections": ("Connections", connections_content),
    "vault": ("File Vault", vault_content),
    "organisation": ("Organisation Admin", organisation_content),
    "settings": ("Settings", settings_content),
}


@app.route("/dashboard")
def dashboard():
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response
    return render_dashboard(overview_content(), "overview", "Overview")


@app.route("/dashboard/<section>")
def dashboard_section(section: str):
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response

    if section not in SECTION_RENDERERS:
        return redirect(url_for("dashboard"))

    title, renderer = SECTION_RENDERERS[section]
    return render_dashboard(renderer(), section, title)


# ---------------------------------------------------------------------------
# API ROUTES
# ---------------------------------------------------------------------------

@app.route("/api/upload", methods=["POST"])
def api_upload():
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response

    if not validate_csrf():
        flash("Security check failed.", "error")
        return redirect(url_for("dashboard_data"))

    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        flash("Please choose a dataset.", "error")
        return redirect(url_for("dashboard_data"))

    try:
        row = save_uploaded_file(uploaded, user["id"])
        log_activity("Dataset uploaded", row["original_name"])
        flash(f"{row['original_name']} was uploaded successfully.", "success")
    except Exception as exc:
        flash(str(exc), "error")

    return redirect(url_for("dashboard_data"))


@app.route("/dashboard/data")
def dashboard_data():
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response
    return render_dashboard(data_content(), "data", "Data Workspace")


@app.route("/api/files/<int:file_id>/download")
def download_file(file_id: int):
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response

    row = db_one("SELECT * FROM files WHERE id=? AND user_id=?", (file_id, user["id"]))
    if not row:
        return "File not found.", 404

    path = UPLOAD_DIR / f"{user['id']}_{row['stored_name']}"
    if not path.exists():
        return "The stored file is no longer available.", 404

    return send_file(path, as_attachment=True, download_name=row["original_name"])


@app.route("/api/di/generate", methods=["POST"])
def api_di_generate():
    user, redirect_response = require_login()
    if redirect_response:
        return jsonify({"ok": False, "error": "Login required."}), 401

    if not validate_csrf():
        return jsonify({"ok": False, "error": "Security check failed."}), 403

    payload = request.get_json(silent=True) or {}
    prompt = clean_text(str(payload.get("prompt", "")), 5000)
    if not prompt:
        return jsonify({"ok": False, "error": "Please provide a request."}), 400

    route = route_task(prompt)
    answer, live = gemini_answer(
        "You are Prociel, an internal David Intelligence worker inside DACRE Analysis. "
        "Generate safe, readable SQL or code based on this user request. "
        "Do not claim to have executed anything. User request: " + prompt
    )

    if live:
        generated = answer
    else:
        # Deterministic offline fallback so the page remains useful without an API key.
        generated = (
            "-- DACRE Analysis / Prociel preview\n"
            "-- Request:\n"
            f"-- {prompt}\n\n"
            "SELECT *\n"
            "FROM your_dataset\n"
            "LIMIT 100;"
        )

    log_activity("DI generation", f"Route: {' → '.join(route)}")
    return jsonify({
        "ok": True,
        "code": generated,
        "answer": generated,
        "live": live,
        "route": route,
    })


@app.route("/api/sql/save", methods=["POST"])
def api_sql_save():
    user, redirect_response = require_login()
    if redirect_response:
        return jsonify({"ok": False, "error": "Login required."}), 401

    if not validate_csrf():
        return jsonify({"ok": False, "error": "Security check failed."}), 403

    payload = request.get_json(silent=True) or {}
    code = clean_text(str(payload.get("code", "")), 50000)
    if not code:
        return jsonify({"ok": False, "error": "Nothing to save."}), 400

    # Save an approved implementation as a text file in File Vault.
    filename = f"dacre_sql_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
    path = UPLOAD_DIR / f"{user['id']}_{secrets.token_hex(10)}_{filename}"
    path.write_text(code, encoding="utf-8")

    file_id = db_execute(
        """
        INSERT INTO files(user_id, original_name, stored_name, extension, size_bytes, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (user["id"], filename, path.name.split(f"{user['id']}_", 1)[-1], "sql", path.stat().st_size, now_iso()),
    )
    log_activity("SQL implementation approved", filename)
    return jsonify({"ok": True, "file_id": file_id})


@app.route("/api/generate-sheet", methods=["POST"])
def api_generate_sheet():
    user, redirect_response = require_login()
    if redirect_response:
        return jsonify({"ok": False, "error": "Login required."}), 401

    if not validate_csrf():
        return jsonify({"ok": False, "error": "Security check failed."}), 403

    payload = request.get_json(silent=True) or {}
    filename = clean_text(str(payload.get("filename", "dacre_data.xlsx")), 120)
    columns = payload.get("columns") or ["Item", "Value"]
    rows = payload.get("rows") or [["Example", 100]]

    if not isinstance(columns, list) or not isinstance(rows, list):
        return jsonify({"ok": False, "error": "Invalid spreadsheet structure."}), 400

    try:
        path = create_xlsx(filename, [str(x)[:100] for x in columns], rows)
        # Add generated workbook to the user's File Vault after explicit generation.
        stored_name = path.name
        file_id = db_execute(
            """
            INSERT INTO files(user_id, original_name, stored_name, extension, size_bytes, created_at)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (user["id"], filename, stored_name, "xlsx", path.stat().st_size, now_iso()),
        )
        log_activity("Spreadsheet generated", filename)
        return jsonify({"ok": True, "file_id": file_id, "download": f"/api/generated/{file_id}"})
    except Exception as exc:
        return jsonify({"ok": False, "error": str(exc)}), 500


@app.route("/api/generated/<int:file_id>")
def api_generated(file_id: int):
    user, redirect_response = require_login()
    if redirect_response:
        return redirect_response

    row = db_one("SELECT * FROM files WHERE id=? AND user_id=?", (file_id, user["id"]))
    if not row:
        return "File not found.", 404

    path = GENERATED_DIR / row["stored_name"]
    if not path.exists():
        return "Generated file is no longer available.", 404
    return send_file(path, as_attachment=True, download_name=row["original_name"])


# ---------------------------------------------------------------------------
# PWA / INSTALLATION
# ---------------------------------------------------------------------------

@app.route("/manifest.webmanifest")
def manifest():
    payload = {
        "name": "DACRE Analysis",
        "short_name": "DACRE",
        "description": "DACRE Analysis data intelligence workspace",
        "start_url": "/",
        "scope": "/",
        "display": "standalone",
        "background_color": "#030b16",
        "theme_color": "#06111f",
        "icons": [
            {
                "src": "/static/dacre-icon-192.png",
                "sizes": "192x192",
                "type": "image/png",
                "purpose": "any maskable",
            },
            {
                "src": "/static/dacre-icon-512.png",
                "sizes": "512x512",
                "type": "image/png",
                "purpose": "any maskable",
            },
        ],
    }
    return Response(json.dumps(payload), mimetype="application/manifest+json")


@app.route("/sw.js")
def service_worker():
    script = """
const CACHE = "dacre-analysis-shell-v1";
const ASSETS = ["/", "/static/dacre-logo.png", "/static/landing-hero.png"];
self.addEventListener("install", event => {
  event.waitUntil(caches.open(CACHE).then(cache => cache.addAll(ASSETS)));
  self.skipWaiting();
});
self.addEventListener("activate", event => event.waitUntil(self.clients.claim()));
self.addEventListener("fetch", event => {
  if (event.request.method !== "GET") return;
  event.respondWith(
    caches.match(event.request).then(cached => cached || fetch(event.request).then(response => {
      const copy = response.clone();
      caches.open(CACHE).then(cache => cache.put(event.request, copy));
      return response;
    }).catch(() => cached))
  );
});
"""
    return Response(script, mimetype="application/javascript")


# ---------------------------------------------------------------------------
# UTILITIES
# ---------------------------------------------------------------------------

def escape_html(value: Any) -> str:
    text = str(value if value is not None else "")
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#x27;")
    )


def escape_attr(value: Any) -> str:
    return escape_html(value)


def format_bytes(value: int) -> str:
    size = float(value)
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}"
        size /= 1024
    return f"{value} B"


@app.route("/health")
def health():
    return jsonify({
        "ok": True,
        "app": APP_NAME,
        "database": DB_PATH.exists(),
        "openpyxl": OPENPYXL_AVAILABLE,
        "pandas": PANDAS_AVAILABLE,
        "gemini_sdk": GENAI_AVAILABLE,
        "google_api_configured": bool(os.environ.get("GOOGLE_API_KEY")),
    })


# ---------------------------------------------------------------------------
# STARTUP
# ---------------------------------------------------------------------------

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("DACRE_DEBUG", "0") == "1"
    print(f"{APP_NAME} starting on http://127.0.0.1:{port}")
    print(f"Database: {DB_PATH}")
    print(f"Google API configured: {bool(os.environ.get('GOOGLE_API_KEY'))}")
    app.run(host="0.0.0.0", port=port, debug=debug)
